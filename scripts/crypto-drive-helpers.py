#!/usr/bin/env python3
"""Helper functions for crypto drive scanning — spreadsheet/doc extraction."""

import json
import sys

CRYPTO_KEYWORDS = [
    "bitcoin", "btc", "ethereum", "eth", "litecoin", "ltc", "dogecoin", "doge",
    "monero", "xmr", "wallet", "private key", "seed", "mnemonic", "passphrase",
    "coinbase", "binance", "kraken", "gemini", "crypto", "blockchain",
    "address", "0x", "xpub", "xprv", "keystore",
]


def scan_spreadsheet(filepath: str) -> str | None:
    """Scan a spreadsheet for crypto-related content. Returns preview or None."""
    ext = filepath.rsplit(".", 1)[-1].lower()

    try:
        if ext == "csv":
            return _scan_csv(filepath)
        elif ext in ("xlsx", "xls"):
            return _scan_xlsx(filepath)
    except Exception as e:
        print(f"Error scanning {filepath}: {e}", file=sys.stderr)

    return None


def _scan_csv(filepath: str) -> str | None:
    import csv

    matches = []
    with open(filepath, "r", errors="ignore") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i > 5000:
                break
            text = " ".join(str(c) for c in row).lower()
            for kw in CRYPTO_KEYWORDS:
                if kw in text:
                    matches.append(f"Row {i + 1}: {' | '.join(str(c) for c in row[:5])}")
                    break
    if matches:
        return "\n".join(matches[:10])
    return None


def _scan_xlsx(filepath: str) -> str | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed", file=sys.stderr)
        return None

    matches = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 5000:
                break
            text = " ".join(str(c) for c in row if c is not None).lower()
            for kw in CRYPTO_KEYWORDS:
                if kw in text:
                    preview = " | ".join(str(c) for c in row[:5] if c is not None)
                    matches.append(f"[{sheet}] Row {i + 1}: {preview}")
                    break
    wb.close()
    if matches:
        return "\n".join(matches[:10])
    return None


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: crypto-drive-helpers.py scan-spreadsheet <file>", file=sys.stderr)
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "scan-spreadsheet":
        result = scan_spreadsheet(sys.argv[2])
        if result:
            print(result)
        else:
            print("null")
    else:
        print(f"Unknown command: {cmd}", file=sys.stderr)
        sys.exit(1)
